"""


pip install openpyxl --upgrade

"""
# for glob stuff
import glob

# for argv and sys.exit
import sys, getopt

# colorize
from colorama import init, Fore, Back, Style
init()


# for spreadsheet stuff
from openpyxl import load_workbook

def list_files(filetype):
    """
    List Files that match
    """
    files = glob.glob("*." + filetype)
    print(Fore.RED + '-'*55)
    print(Fore.RED + "Compatible Files Found:")
    print(Fore.RED + '-'*55)
    for f in files:
        print(Fore.GREEN + f)


def get_col(sheet,col):
    """
    dump the column we are looking for
    """
    print(Fore.RED + '-'*55)
    print(Fore.RED + "DUMPING COL " + str(col))
    print(Fore.RED + '-'*55 + Style.RESET_ALL)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        print(row[col])



def count_cols(sheet):
    """
    Given a sheet, tell me the column that have stuff
    Immediately returns at first blank header
    """
    cols = 0
    for col in sheet.iter_cols(min_col=1,max_col=20,min_row=1,max_row=2):
    #bug https://bitbucket.org/openpyxl/openpyxl/issues/514/cell-max_row-reports-higher-than-actual
        for cell in col:
            if cell.value is None:
                return cols
            else:
                cols = cols + 1
        

def dump_header(sheet,cols):
    """
    Given a sheet, and colums, tell me the header names and first row values
    Uses count_cols and doesn't do more than it returns
    """

    # build our data first so we can format it together later
    row1 = []
    row2 = []

    # get the header
    for row in sheet.iter_rows(min_row=1, max_col=cols, max_row=1):
        for cell in row: row1.append(cell.value)

    # get first row
    for row in sheet.iter_rows(min_row=2, max_col=cols, max_row=2):
        for cell in row: row2.append(cell.value)

    # print header
    print(Fore.RED + '-'*55)
    print(Fore.RED + f"{'#':>4}",end="\t")
    print(Fore.RED + f"{'HEADER':>22}",end="\t")
    print(Fore.RED + f"{'DATA'}")
    print(Fore.RED + '-'*55)

    # print them out left and right justified and numbered for first row
    for i in range(0,cols):
        row1_data = str(row1[i])
        row2_data = str(row2[i])
        print(Fore.YELLOW + f"{i:>4}",end="\t")
        print(Fore.GREEN + f"{row1_data:>22}",end="\t")
        print(Fore.YELLOW + f"{row2_data}")


def help_message(msg):
    """
    Shows help message
    """
    print(Fore.RED + '-'*55)
    print(Fore.RED + msg)
    print(Fore.RED + '-'*55 + Fore.GREEN)
    print(Style.BRIGHT + Fore.YELLOW + "(Argument Order is Specific!!)")
    print(Style.RESET_ALL + "Examples:")
    print(sys.argv[0] + " -h")
    print(sys.argv[0] + " -i <inputfile> -p")
    print(sys.argv[0] + " -i <inputfile> -d <column>\n")
    print("-h prints this help message")
    print("-i inputfile is your source, input file")
    #print("-o outputfile is your source, output file")
    print("-d is to dump a column, you must specify which column")
    print("-p is to peek at the header and first row")
    print(Style.RESET_ALL)
    list_files('xlsx')
    sys.exit(2)

def is_accessible(path, mode='r'):
    """
    Check if the file or directory at `path` can
    be accessed by the program using `mode` open flags.
    """
    try:
        f = open(path, mode)
        f.close()
    except IOError:
        return False
    return True

def main():

    workbook = ''
    sheetnames = ''
    sheet = ''
    inputfile = ''
    i = 0

    # basic opt verification
    try:
        opts, args = getopt.getopt(sys.argv[1:],"h?pd:i:o:", ["help", '?'])
    except getopt.GetoptError as err:
        help_message(err)



    for opt, arg in opts:
        if opt in ("-h", "-help", '-?'):
            help_message("INSTRUCTIONS")

        if opt in ("-i"):
            i = 1

            # because they could say -i but not give a command
            if len(sys.argv) < 4:
                help_message("ADDITIONAL ARGUMENTS REQUIRED")

            # verify exists and is readable
            if (is_accessible(arg)):
                inputfile = arg
                # load a workbook object
                workbook = load_workbook(filename=inputfile, data_only=True)

                # get the sheet names
                sheetnames = workbook.sheetnames
                sheet = workbook.active
            else: help_message("CANNOT READ FILE")

        if opt in ("-p"):
            if i:
                cols = count_cols(sheet)
                dump_header(sheet,cols)
                sys.exit(2)
            else:
                help_message("MISSING INPUT FILE")


        if opt in ("-d"):
            if i:
                get_col(sheet,int(arg))
                sys.exit()
            else:
                help_message("MISSING INPUT FILE")
        #else:
        #    help_message("INVALID OPTION")

    sys.exit()



if __name__ == "__main__":
    if len(sys.argv) > 1:
        main()
    else:
        help_message("INSTRUCTIONS")
